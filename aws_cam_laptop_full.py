# -*- coding: utf-8 -*-
import boto3
import io
from PIL import Image
import time
import cv2
import openpyxl
from gtts import gTTS
from playsound import playsound
import openpyxl as xl
from pathlib import Path

def text2speech():
    # Tạo danh sách tên theo ID tương ứng
    list_face = ['Tổng Bí Thư',
                'Chủ Tịch Nước',
                'Thủ Tướng',
                'Chủ Tịch Quốc Hội',
                'Võ Văn Thưởng',
                'Trương Thị Mai',
                'Phạm Bình Minh',
                'Nguyễn Văn Nên',
                'Tô Lâm',
                'Phan Đình Trạc',
                'Trần Thanh Mẫn',
                'Trần Tuấn Anh',
                'Trần Cẩm Tú',
                'Phan Văn Giang',
                'Nguyễn Hòa Bình',
                'Nguyễn Xuân Thắng',
                'Lương Cường',
                'Trần Tuấn Anh',
                'Đinh Tiến Dũng',
                'Vũ Đức Đam',
                'Thống Đốc Ngân Hàng',
                'Trần Sỹ Thanh',
                'Hồ Hùng Anh',
                'Đỗ Quang Hiển',
                'Đỗ Minh Phú',
                'Thái Hương',
                'Võ Quốc Thắng',
                'Nguyễn Thị Nga',
                'Đỗ Anh Tuấn', 
                'Tổng Giám Đốc',
                'Chị Ngoãn',
                'Trung Hiếu',
                'Quốc Anh',
                'Văn Thành',
                'Thanh Phong',
                'Hiếu T2', 
                'Phú',
                'Huyền',
                ]
    # Mở wb đã lưu ID nhận dạng được
    wb = xl.load_workbook('./excelCreate.xlsx')
    activeSheet = wb.active
    sheet = wb['Sheet']
    list_1 = list(sheet.columns)[0]
    cell = 0
    face = []
    for cellObj in list_1:
        # Chỉ đến cột A và thêm giá trị vào face[]
        cell_Ai = sheet["A{}".format(cell+1)]
        # print(type(int( cell_Ai.value)))
        add = int(cell_Ai.value)
        face.append(add)
        cell +=1
        # print(face)
    # Sắp xếp ID để nhận diện khách VIP
    lenth = len(face)
    for i in range(0,lenth-1):
        for j in range(i+1,lenth):
            if (face[i]>face[j]):
                # Hoán đổi vị trí
                tmp = face[i]
                face[i] = face[j]
                face[j] = tmp
    # Trả về speech
    for ID in face:         
        speech = list_face[ID]
        text = 'Xin chào: ' + speech
        output = gTTS(text, lang="vi", slow=False)
        output.save("output.mp3")
        audio = Path().cwd() / "output.mp3"
        playsound(audio)
        # playsound.playsound('output.mp3', True)
        print(speech)
    return list_face

def face_reco():
    person = None
    t1 = time.time()
    # Tạo workbook để lưu file excel
    wb = openpyxl.Workbook()
    activesheet = wb.active
    activesheet.title ="Sheet"
    wb.create_sheet()
    sheet = wb['Sheet']
    # Kết nối đến sever AWS
    rekognition = boto3.client('rekognition', region_name='ap-southeast-1')
    dynamodb = boto3.client('dynamodb', region_name='ap-southeast-1')
    t2 = time.time()
    # Lấy ảnh từ thu mục và chuyển về ảnh nhị phân
    img_name = "./image_face_0.jpg"
    # image_name_1 = cv2.imread(img_name)
    image = Image.open(img_name)
    stream = io.BytesIO()
    image.save(stream,format="JPEG")
    image_binary = stream.getvalue()
    # cam = cv2.VideoCapture(0)
    # frame = cam.read()
    # img_name = frame
    # image = Image.open(img_name)
    # stream = io.BytesIO()
    # image.save(stream,format="JPG")
    # image_binary = stream.getvalue()
    # Detect face by rekognition
    response = rekognition.detect_faces(Image={'Bytes':image_binary})
    all_faces = response['FaceDetails']
    t3 = time.time()
    if all_faces == []:
        return
    else:
        # Tạo list khuôn mặt
        boxes = []
        # Lấy size ảnh
        image_width = image.size[0]
        image_height = image.size[1]
        # Cắt mặt detect được
        for face in all_faces:
            box = face['BoundingBox']
            x1 = int(box['Left'] * image_width) * 0.9
            y1 = int(box['Top'] * image_height) * 0.9
            x2 = int(box['Left'] * image_width + box['Width'] * image_width) * 1.10
            y2 = int(box['Top'] * image_height + box['Height']  * image_height) * 1.10
            image_crop = image.crop((x1,y1,x2,y2))
            # Chuyển về ảnh nhị phân
            stream = io.BytesIO()
            image_crop.save(stream,format="JPEG")
            image_crop_binary = stream.getvalue()
            # Gửi hình ảnh được cắt riêng lẻ lên Amazon Rekognition
            response = rekognition.search_faces_by_image(
                    FaceMatchThreshold=98,
                    MaxFaces=10,
                    CollectionId='faceRecoSingv1',
                    Image={'Bytes':image_crop_binary}                                       
                    )
            if len(response['FaceMatches']) > 0:
                # Return results
                for match in response['FaceMatches']:
                    face = dynamodb.get_item(
                        TableName='faceRecoSingt1',               
                        Key={'RekognitionId': {'S': match['Face']['FaceId']}})
                    if 'Item' in face:
                        person = face['Item']['FullName']['S']
                        # Save person-reco
                        numFace=1
                        if person != sheet["A{}".format(numFace)]:
                            sheet["A{}".format(numFace)] = person
                            sheet["A{}".format(numFace)].value
                            wb.save('excelCreate.xlsx')
                            numFace += 1
                        else:
                            sheet["A{}".format(numFace)] = person
                            sheet["A{}".format(numFace)].value
                            wb.save('excelCreate.xlsx')
                    else:
                        person = 'no match found'
        t4 = time.time()               
        print("Thời gian nhận dạng mặt: ",t4-t1)
        text2speech()
    
if __name__ == "__main__":
    face_reco()
        # k = cv2.waitKey(1)
        # if k%256 == 27:
        # # ESC pressed
        #     print("closing")
        #     break




